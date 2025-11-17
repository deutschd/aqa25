
import re
from datetime import datetime
from pathlib import Path
from typing import List, Dict, Optional

import pandas as pd

# ----------- НАСТРОЙКИ ------------
PDF_PATH = Path(r"D:\EXPERIMENT\test.pdf")
OUT_PATH = Path(r"D:\EXPERIMENT\outcome.xlsx")
# ----------------------------------


# ---------- Извлечение текста из PDF ----------
def extract_text_from_pdf(pdf_path: Path) -> str:
    text = ""
    # 1) Быстрый вариант: PyPDF2
    try:
        import PyPDF2
        with open(pdf_path, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                page_text = page.extract_text() or ""
                text += page_text + "\n"
    except Exception:
        pass

    # 2) Точный вариант: pdfminer.six
    if not text.strip():
        try:
            from pdfminer.high_level import extract_text as pdfminer_extract
            text = pdfminer_extract(pdf_path.as_posix())
        except Exception:
            pass

    if not text.strip():
        raise RuntimeError(
            "Не удалось извлечь текст из PDF. Проверь файл или сохрани его в текстовом виде."
        )
    return text


# ---------- Разбивка на кейсы ----------
def split_into_blocks(text: str) -> List[str]:
    """
    Ищем строки, начинающиеся с '#<digits>' или '# <digits>'.
    Собираем блок от этой строки до следующего заголовка.
    """
    text = text.replace("\r", "\n")
    text = re.sub(r"[ \t]+\n", "\n", text)
    text = re.sub(r"\n{3,}", "\n\n", text)

    lines = text.splitlines()
    indices = [i for i, line in enumerate(lines) if re.match(r"^#\s*\d{3,}", line.strip())]

    blocks: List[str] = []
    if indices:
        for idx, start in enumerate(indices):
            end = indices[idx + 1] if idx + 1 < len(indices) else len(lines)
            block = "\n".join(lines[start:end]).strip()
            if block:
                blocks.append(block)
    else:
        # если совсем нет маркеров, вернём один общий блок
        return [text]

    return blocks


# ---------- Вспомогательные парсеры ----------
STOP_LABELS = [
    "Tags", "Custom fields", "Custom Fields", "Members", "Description",
    "Precondition", "Prerequisite", "Scenario", "Steps", "Expected Result", "Expected result"
]

def _is_stop(line: str) -> bool:
    return any(re.match(rf"^\s*{lab}\s*:", line, flags=re.IGNORECASE) for lab in STOP_LABELS)

def _capture_section(whole: str, label_variants, stop_labels) -> Optional[str]:
    pattern = r"(?:" + "|".join([re.escape(v) for v in label_variants]) + r")\s*:\s*"
    starts = list(re.finditer(pattern, whole, re.IGNORECASE))
    if not starts:
        return None
    start = starts[0].end()

    stop_idx = len(whole)
    for sl in stop_labels:
        m2 = re.search(r"\n\s*" + re.escape(sl) + r"\s*:", whole[start:], re.IGNORECASE)
        if m2:
            stop_idx = start + m2.start()
            break

    return whole[start:stop_idx].strip()


def _parse_expected_fields(exp: Optional[str]) -> Dict[str, Optional[str]]:
    """
    Извлекаем полезные поля из Expected Result (если там JSON-подобный блок):
      ERROR_NOTE, ERROR_CODE, TRANS_ID, OPER_DATE, GROUP_ID
    """
    if not exp:
        return {"ERROR_NOTE": None, "ERROR_CODE": None, "TRANS_ID": None, "OPER_DATE": None, "GROUP_ID": None}

    def rex(pattern, flags=re.IGNORECASE):
        m = re.search(pattern, exp, flags)
        return m.group(1).strip() if m else None

    # Ловим и со/без кавычек, и со/без пробелов
    error_note = rex(r'ERROR_NOTE"\s*:\s*"([^"]+)"')
    if not error_note:
        error_note = rex(r"ERROR_NOTE'\s*:\s*'([^']+)'")
    error_code = rex(r'ERROR_CODE"\s*:\s*([0-9\-]+)')
    trans_id   = rex(r'TRANS_ID"\s*:\s*([0-9\-]+)')
    oper_date  = rex(r'OPER_DATE"\s*:\s*"([^"]+)"')
    group_id   = rex(r'GROUP_ID"\s*:\s*([0-9\-]+)')

    return {
        "ERROR_NOTE": error_note,
        "ERROR_CODE": error_code,
        "TRANS_ID": trans_id,
        "OPER_DATE": oper_date,
        "GROUP_ID": group_id,
    }


# ---------- Парсинг одного кейса ----------
def parse_case_block(case_text: str) -> Dict[str, Optional[str]]:
    d: Dict[str, Optional[str]] = {
        "ID": None,
        "Case Name": None,
        "Status/State": None,
        "Suite": None,
        "Tags": None,
        "Custom Fields": None,
        "Members": None,
        "Description": None,
        "Precondition": None,
        "Scenario": None,
        "Expected Result": None,
        "Endpoint/Host": None,
        "HTTP Request": None,
        "HTTP Response (expected)": None,
        "Created At": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Source Page Hint": None,
    }

    # --- заголовок (#ID + Name + [Status])
    lines = [ln for ln in case_text.splitlines() if ln.strip()]
    if not lines:
        return d

    # Найти строку с #ID
    id_idx = None
    for i, ln in enumerate(lines):
        if re.search(r"^#\s*\d{3,}", ln.strip()):
            id_idx = i
            break

    header_chunk = lines[0] if id_idx is None else lines[id_idx]

    # Склеить до 3 строк хедера, пока не встретим стоп-секцию
    j = (id_idx or 0) + 1
    while j < len(lines) and not _is_stop(lines[j]) and j <= (id_idx or 0) + 3:
        header_chunk += " " + lines[j].strip()
        j += 1
    header_chunk = re.sub(r"\s+", " ", header_chunk).strip()

    # Спарсить ID / Name / [Status]
    m = re.search(r"#\s*(\d{3,})\s+(.+?)(?:\s*\[([^\]]+)\])?$", header_chunk)
    if m:
        d["ID"] = m.group(1)
        d["Case Name"] = m.group(2).strip()
        d["Status/State"] = (m.group(3) or "").strip() if m.group(3) else None
    else:
        mm = re.search(r"#\s*(\d{3,})", header_chunk)
        if mm:
            d["ID"] = mm.group(1)
            tail = header_chunk[mm.end():].strip(" -—:")
            d["Case Name"] = tail or None

    whole = case_text

    # Секции
    d["Tags"] = _capture_section(
        whole, ["Tags"],
        ["Custom fields", "Custom Fields", "Members", "Description", "Precondition", "Prerequisite", "Scenario", "Steps", "Expected Result", "Expected result"],
    )

    cf = _capture_section(
        whole, ["Custom fields", "Custom Fields"],
        ["Members", "Description", "Precondition", "Prerequisite", "Scenario", "Steps", "Expected Result", "Expected result"],
    )
    d["Custom Fields"] = cf
    if cf:
        m_suite = re.search(r"Suite\s*:\s*([^\n]+)", cf, re.IGNORECASE)
        if m_suite:
            d["Suite"] = m_suite.group(1).strip()

    d["Members"] = _capture_section(
        whole, ["Members"],
        ["Description", "Precondition", "Prerequisite", "Scenario", "Steps", "Expected Result", "Expected result"],
    )

    d["Description"] = _capture_section(
        whole, ["Description"],
        ["Precondition", "Prerequisite", "Scenario", "Steps", "Expected Result", "Expected result"],
    )

    d["Precondition"] = _capture_section(
        whole, ["Precondition", "Prerequisite"],
        ["Scenario", "Steps", "Expected Result", "Expected result"],
    )

    scen = _capture_section(
        whole, ["Scenario", "Steps"],
        ["Expected Result", "Expected result"],
    )
    d["Scenario"] = scen

    if scen:
        # endpoint + host
        m_post = re.search(r">\s*(GET|POST|PUT|PATCH|DELETE)\s+([^\s>]+)", scen, re.IGNORECASE)
        if m_post:
            d["Endpoint/Host"] = m_post.group(2).strip()
        m_host = re.search(r"> *Host:\s*([^\s>]+)", scen, re.IGNORECASE)
        if m_host:
            d["Endpoint/Host"] = f"{d['Endpoint/Host'] or ''} (Host: {m_host.group(1)})".strip()
        d["HTTP Request"] = scen

    exp = _capture_section(whole, ["Expected Result", "Expected result"], []) or None
    d["Expected Result"] = exp
    d["HTTP Response (expected)"] = exp

    return d


# ---------- Основной сценарий ----------
def main():
    # проверим пути
    if not PDF_PATH.exists():
        raise FileNotFoundError(f"PDF не найден: {PDF_PATH}")
    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    print(f"[i] Читаю PDF: {PDF_PATH}")
    raw_text = extract_text_from_pdf(PDF_PATH)

    print("[i] Разбиваю на кейсы...")
    blocks = split_into_blocks(raw_text)
    print(f"[i] Найдено блоков: {len(blocks)}")

    rows = [parse_case_block(b) for b in blocks]
    # убираем полностью пустые
    rows = [r for r in rows if any(v for k, v in r.items() if k not in ("Created At", "Source Page Hint"))]

    if not rows:
        raise RuntimeError("Кейсы не распознаны. Убедись, что каждый начинается с строки вида: #18369 Название ...")

    # DataFrame и доп. поля из Expected Result
    df = pd.DataFrame(rows)

    # Гарантируем порядок колонок
    columns = [
        "ID","Case Name","Status/State","Suite","Tags","Custom Fields","Members",
        "Description","Precondition","Scenario","Expected Result",
        "Endpoint/Host","HTTP Request","HTTP Response (expected)",
        "Created At","Source Page Hint"
    ]
    for c in columns:
        if c not in df.columns:
            df[c] = None
    df = df[columns]

    # Добавим разбор Expected Result в отдельные колонки
    extra_cols = ["ERROR_NOTE", "ERROR_CODE", "TRANS_ID", "OPER_DATE", "GROUP_ID"]
    extra_data = df["Expected Result"].apply(_parse_expected_fields).apply(pd.Series)
    for c in extra_cols:
        if c not in extra_data.columns:
            extra_data[c] = None
    df = pd.concat([df, extra_data[extra_cols]], axis=1)

    print(f"[i] Сохраняю Excel: {OUT_PATH}")
    df.to_excel(OUT_PATH.as_posix(), index=False, sheet_name="Cases")

    # Стилизация Excel через openpyxl
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        wb = load_workbook(OUT_PATH.as_posix())
        ws = wb.active  # "Cases"

        # автофильтр и заморозка шапки
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}1"
        ws.freeze_panes = "A2"

        # стили
        header_fill = PatternFill(start_color="FFE8F1FF", end_color="FFE8F1FF", fill_type="solid")
        header_font = Font(bold=True)
        wrap = Alignment(wrap_text=True, vertical="top")
        thin = Side(style="thin", color="DDDDDD")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center")

        # разумные ширины
        width_map = {
            "A": 10,  # ID
            "B": 52,  # Case Name
            "C": 16,  # Status/State
            "D": 34,  # Suite
            "E": 26,  # Tags
            "F": 42,  # Custom Fields
            "G": 28,  # Members
            "H": 60,  # Description
            "I": 42,  # Precondition
            "J": 60,  # Scenario
            "K": 60,  # Expected Result
            "L": 40,  # Endpoint/Host
            "M": 60,  # HTTP Request
            "N": 60,  # HTTP Response (expected)
            "O": 20,  # Created At
            "P": 18,  # Source Page Hint
            "Q": 26,  # ERROR_NOTE
            "R": 14,  # ERROR_CODE
            "S": 22,  # TRANS_ID
            "T": 24,  # OPER_DATE
            "U": 14,  # GROUP_ID
        }

        # применим ширины и wrap+бордеры
        for col_letter, width in width_map.items():
            if col_letter <= get_column_letter(ws.max_column):
                ws.column_dimensions[col_letter].width = width

        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = wrap
                cell.border = border

        wb.save(OUT_PATH.as_posix())
    except Exception as e:
        print(f"[!] Не удалось применить стили к Excel: {e}")

    print("[✓] Готово!")
    print(f"    Файл: {OUT_PATH}")
    print(f"    Строк: {len(df)}")


if __name__ == "__main__":
    main()
